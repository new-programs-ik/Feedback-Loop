"""
flagship_import.py — import the confidential Flagship ML workbook into Supabase.

The .xlsx is read locally and NEVER committed (it's gitignored). Parses each cohort tab
(Date · Topic · Instructor(live) · Thursday review · Wednesday coaching) and upserts the
course, cohorts, instructors and per-cohort class schedule.

  python flagship_import.py --dry-run   # parse + summarize only, no DB writes
  python flagship_import.py             # import into DB (DATABASE_URL from .env)
"""
from __future__ import annotations

import datetime
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "Machine Learning Program with Agentic AI _ Aug 2025 Onwards (1).xlsx")
COURSE = "Flagship ML"
COURSE_FULL = "Machine Learning Program with Agentic AI"


def clean(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


_JUNK = {"#ref!", "na", "n/a", "tbd", "tba", "-", "?", "none", "break"}


def instructor(v) -> str:
    """Return a clean instructor name, or '' for blanks / broken refs / notes."""
    n = clean(v)
    low = n.lower()
    if not n or low in _JUNK:
        return ""
    if "same as" in low or "from here" in low or "onwards" in low:
        return ""
    if not re.search(r"[A-Za-z]", n):  # must contain a letter
        return ""
    return n


def is_cohort_tab(title: str) -> bool:
    t = title.lower()
    return "cohort" in t and "resou" not in t and "uplevel" not in t


def parse_tab(ws) -> list[dict]:
    header_row, headers = None, None
    for r in range(1, 16):
        vals = [clean(ws.cell(r, c).value).lower() for c in range(1, (ws.max_column or 1) + 1)]
        if any("topic" in v for v in vals) and any("date" in v for v in vals):
            header_row, headers = r, vals
            break
    if header_row is None:
        return []

    def find(pred):
        for c, v in enumerate(headers, start=1):
            if pred(v):
                return c
        return None

    date_c = find(lambda v: "date" in v)
    topic_c = find(lambda v: "topic" in v)
    live_c = find(lambda v: v == "instructor" or ("instructor" in v and not any(
        x in v for x in ("thursday", "wednesday", "review", "coaching"))))
    thu_c = find(lambda v: "thursday" in v or "review" in v)
    wed_c = find(lambda v: "wednesday" in v or "coaching" in v)

    out, week = [], 0
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        topic = clean(ws.cell(r, topic_c).value) if topic_c else ""
        date = as_date(ws.cell(r, date_c).value) if date_c else None
        if not topic or topic.lower() == "break" or not date:
            continue
        week += 1
        out.append({
            "week": week, "date": date, "topic": topic,
            "live": instructor(ws.cell(r, live_c).value) if live_c else "",
            "review": instructor(ws.cell(r, thu_c).value) if thu_c else "",
            "coaching": instructor(ws.cell(r, wed_c).value) if wed_c else "",
        })
    return out


def parse() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    cohorts = []
    for ws in wb.worksheets:
        if is_cohort_tab(ws.title):
            classes = parse_tab(ws)
            if classes:
                cohorts.append({"name": clean(ws.title), "classes": classes})
    return cohorts


def summarize(cohorts: list[dict]) -> None:
    instructors, total, unassigned = set(), 0, 0
    for co in cohorts:
        for cl in co["classes"]:
            total += 1
            for k in ("live", "review", "coaching"):
                if cl[k]:
                    instructors.add(cl[k])
            if not cl["live"]:
                unassigned += 1
    print(f"cohorts: {len(cohorts)} | classes: {total} | distinct instructors: {len(instructors)} | "
          f"classes without a live instructor: {unassigned}")
    for co in cohorts:
        ds = [c["date"] for c in co["classes"]]
        print(f"  {co['name']:<30} {len(co['classes']):>2} classes  {min(ds)} .. {max(ds)}")
    print("instructors:", ", ".join(sorted(instructors)))


def seed(cohorts: list[dict]) -> None:
    import psycopg2
    url = re.search(r"DATABASE_URL=(.+)", open(os.path.join(HERE, ".env")).read()).group(1).strip()
    url = url.replace("postgresql+psycopg2://", "postgresql://")
    conn = psycopg2.connect(url, connect_timeout=15)
    cur = conn.cursor()

    cur.execute("insert into courses(name, slug, description) values (%s,%s,%s) "
                "on conflict (name) do update set description=excluded.description returning id",
                (COURSE, "flagship-ml", COURSE_FULL))
    course_id = cur.fetchone()[0]

    names = {cl[k] for co in cohorts for cl in co["classes"] for k in ("live", "review", "coaching") if cl[k]}
    id_by_name = {}
    for n in sorted(names):
        cur.execute("insert into instructors(name) values (%s) on conflict (name) do update set name=excluded.name returning id", (n,))
        id_by_name[n] = cur.fetchone()[0]

    nclasses = 0
    for co in cohorts:
        ds = [c["date"] for c in co["classes"]]
        cur.execute("insert into cohorts(course_id, name, start_date, end_date) values (%s,%s,%s,%s) "
                    "on conflict (course_id, name) do update set start_date=excluded.start_date, "
                    "end_date=excluded.end_date returning id", (course_id, co["name"], min(ds), max(ds)))
        cohort_id = cur.fetchone()[0]
        for cl in co["classes"]:
            cur.execute(
                "insert into cohort_classes(cohort_id, week_no, class_date, topic, instructor_id, "
                "review_instructor_id, coaching_instructor_id) values (%s,%s,%s,%s,%s,%s,%s) "
                "on conflict (cohort_id, class_date, topic) do update set week_no=excluded.week_no, "
                "instructor_id=excluded.instructor_id, review_instructor_id=excluded.review_instructor_id, "
                "coaching_instructor_id=excluded.coaching_instructor_id, updated_at=now()",
                (cohort_id, cl["week"], cl["date"], cl["topic"],
                 id_by_name.get(cl["live"]), id_by_name.get(cl["review"]), id_by_name.get(cl["coaching"])))
            nclasses += 1

    conn.commit()
    conn.close()
    print(f"seeded: course '{COURSE}' + {len(cohorts)} cohorts + {len(names)} instructors + {nclasses} classes")


if __name__ == "__main__":
    data = parse()
    summarize(data)
    if "--dry-run" not in sys.argv:
        seed(data)
